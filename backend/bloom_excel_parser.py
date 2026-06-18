"""
블룸 엑셀 자동 파싱 모듈
- 실적 엑셀 (PO/생산/출하)
- BOP NCT (자재 부족)
- KPE NCT (자재 부족)
3개 파일을 받아 블룸 카드 노트를 자동 생성한다.

파일명에 의존하지 않고, 시트명 + 내용으로 자동 분류한다.
"""

import openpyxl
from pathlib import Path
from collections import Counter
from typing import Optional, Dict, List, Any


# ========================================
# 1. 파일 자동 분류 (파일명 무시, 내용으로 판별)
# ========================================
def classify_excel(path: str) -> str:
    """
    엑셀 파일을 열어서 종류를 판별한다.
    리턴값: 'po' | 'bop_nct' | 'kpe_nct' | 'unknown'
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheets = set(wb.sheetnames)
    except Exception:
        return 'unknown'

    # 실적 파일 판별
    po_keys = {'출하계획', '날짜별 생산실적'}
    if po_keys.issubset(sheets):
        return 'po'

    # NCT 파일 판별 (일일 수량 요약 시트 + Model 컬럼 값)
    if '일일 수량 요약' in sheets:
        try:
            ws = wb['일일 수량 요약']
            models = []
            for r in ws.iter_rows(min_row=7, max_row=min(ws.max_row, 200), values_only=True):
                if len(r) >= 3 and r[2]:
                    models.append(str(r[2]).strip().upper())
            cnt = Counter(models)
            if not cnt:
                return 'unknown'
            top_model, _ = cnt.most_common(1)[0]
            if 'BOP' in top_model:
                return 'bop_nct'
            if 'CORVA' in top_model or 'KPE' in top_model:
                return 'kpe_nct'
        except Exception:
            pass

    return 'unknown'


# ========================================
# 2. 실적 엑셀 파싱
# ========================================
def parse_po_excel(path: str) -> Dict[str, Dict[str, Any]]:
    """
    실적 엑셀에서 제품별 PO/출하/생산 정보 추출.
    리턴: { '제품명': {'po': N, 'ship_done': N, 'ship_left': N, 'prod': N} }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    result: Dict[str, Dict[str, Any]] = {}

    # --- 출하계획 시트: 총PO / 출하완료 / 출하잔량 ---
    if '출하계획' in wb.sheetnames:
        ws = wb['출하계획']
        for r in ws.iter_rows(min_row=4, values_only=True):
            name = r[1]
            if not name or not isinstance(name, str):
                continue
            name = name.strip()
            if name in ('제품명',):
                continue
            po, ship_done, ship_left = r[2], r[3], r[4]
            if isinstance(po, (int, float)):
                result.setdefault(name, {})
                result[name]['po'] = int(po)
                result[name]['ship_done'] = int(ship_done) if isinstance(ship_done, (int, float)) else 0
                result[name]['ship_left'] = int(ship_left) if isinstance(ship_left, (int, float)) else 0

    # --- 날짜별 생산실적 시트: 생산완료 누적 ---
    if '날짜별 생산실적' in wb.sheetnames:
        ws = wb['날짜별 생산실적']
        for r in ws.iter_rows(min_row=4, values_only=True):
            name_raw = r[1]
            if not name_raw or not isinstance(name_raw, str):
                continue
            # 줄바꿈 / 변경 표기 정리
            name_clean = name_raw.replace('\n', ' ').strip()
            if 'KPE EFM' in name_clean:
                name_clean = 'KPE EFM'
            else:
                name_clean = name_clean.split('생산')[0].strip()

            if isinstance(r[5], (int, float)):
                # 매칭 시도
                matched = None
                for k in result.keys():
                    if k == name_clean or k.replace(' ', '') == name_clean.replace(' ', ''):
                        matched = k
                        break
                if matched is None:
                    # 새 제품 (출하계획에 없는 경우)
                    result[name_clean] = {}
                    matched = name_clean
                result[matched]['prod'] = int(r[5])

    # 기본값 채우기
    for k, v in result.items():
        v.setdefault('po', 0)
        v.setdefault('ship_done', 0)
        v.setdefault('ship_left', 0)
        v.setdefault('prod', 0)

    return result


# ========================================
# 3. NCT 엑셀 파싱
# ========================================
def parse_nct_excel(path: str) -> Dict[str, Any]:
    """
    NCT 엑셀에서 자재 부족 정보 추출.
    리턴: {
        'model': 'BOP' or 'CORVA',
        'total': N,
        'shortage_count': N,
        'top3': [{'name': str, 'future': N}, ...],
        'all_shortages': [...]
    }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if '일일 수량 요약' not in wb.sheetnames:
        return {'model': None, 'total': 0, 'shortage_count': 0, 'top3': [], 'all_shortages': []}

    ws = wb['일일 수량 요약']
    items: List[Dict[str, Any]] = []
    models_seen: List[str] = []

    for r in ws.iter_rows(min_row=7, values_only=True):
        if len(r) < 6:
            continue
        no, name, model, bom, today, future = r[:6]
        if not isinstance(no, (int, float)):
            continue
        if not name:
            continue
        if model:
            models_seen.append(str(model).strip().upper())
        items.append({
            'name': str(name).strip(),
            'model': str(model).strip() if model else '',
            'future': future if isinstance(future, (int, float)) else 0,
        })

    # 대표 Model
    top_model = Counter(models_seen).most_common(1)[0][0] if models_seen else None

    shortages = [x for x in items if x['future'] < 0]
    shortages.sort(key=lambda x: x['future'])

    return {
        'model': top_model,
        'total': len(items),
        'shortage_count': len(shortages),
        'top3': shortages[:3],
        'all_shortages': shortages,
    }


# ========================================
# 4. 노트 자동 생성
# ========================================
DEFAULT_PRODUCT_ORDER = ['YFP', 'KPE CORVA', 'KPE EFM', 'BOP Assy', 'SL7', 'SS8']

# 제품 → NCT Model 매핑
PRODUCT_NCT_MAP = {
    'BOP Assy': 'BOP',
    'KPE CORVA': 'CORVA',
}


def fmt(n) -> str:
    """숫자 포맷팅 (1,234 형태)"""
    try:
        return f'{int(n):,}'
    except Exception:
        return str(n)


def build_bloom_note(
    po_data: Dict[str, Dict[str, Any]],
    bop_nct: Optional[Dict[str, Any]] = None,
    kpe_nct: Optional[Dict[str, Any]] = None,
    base_date: str = '',
) -> Dict[str, Any]:
    """
    파싱된 데이터로 블룸 카드 노트 자동 생성.
    리턴: {
        'title': '<블룸>',
        'sections': [...],
        'raw_text': '...',
        'status_hint': 'RED' | 'ORANGE' | ...,
    }
    """
    # NCT 매핑
    material_map = {}
    if bop_nct and bop_nct.get('model'):
        material_map['BOP Assy'] = bop_nct
    if kpe_nct and kpe_nct.get('model'):
        material_map['KPE CORVA'] = kpe_nct

    products_in_data = list(po_data.keys())
    products = [p for p in DEFAULT_PRODUCT_ORDER if p in po_data]
    # 추가 제품 (default 리스트에 없는)
    for p in products_in_data:
        if p not in products:
            products.append(p)

    sections = []
    has_red = False
    has_orange = False

    # 제품별 진행 현황 섹션
    for i, p in enumerate(products, 1):
        d = po_data.get(p, {})
        po = d.get('po', 0)
        prod = d.get('prod', 0)
        ship_done = d.get('ship_done', 0)
        ship_left = d.get('ship_left', 0)

        pct_prod = (prod / po * 100) if po else 0
        pct_ship = (ship_done / po * 100) if po else 0

        block_lines = [f'{i}) {p}']

        # 원자재
        nct = material_map.get(p)
        if nct and nct['shortage_count'] > 0:
            sc = nct['shortage_count']
            total = nct['total']
            if sc >= 4:
                block_lines.append(f'   [원자재] 🔴 NCT 자재 부족 {sc}/{total}종 (생산 차질 우려)')
                has_red = True
            else:
                block_lines.append(f'   [원자재] 🟡 NCT 자재 부족 {sc}/{total}종')
                has_orange = True
            for it in nct['top3']:
                block_lines.append(f'            · {it["name"]} : {fmt(it["future"])}')
            if sc > 3:
                block_lines.append(f'            ...외 {sc - 3}종')
        elif po == 0:
            block_lines.append(f'   [원자재] -')
        elif po > 0 and prod == 0 and ship_done == 0:
            block_lines.append(f'   [원자재] 미발주')
        else:
            block_lines.append(f'   [원자재] 자재 점검 대상 없음')

        # 입고 (현재 데이터 없음)
        block_lines.append(f'   [입고]   -')

        # 생산
        if po:
            block_lines.append(f'   [생산]   {fmt(prod)} / {fmt(po)} 조립완료 ({pct_prod:.1f}%)')
        else:
            block_lines.append(f'   [생산]   {fmt(prod)} (총 PO 정보 없음)')

        # 납기
        if po:
            block_lines.append(
                f'   [납기]   총 PO {fmt(po)} / 출하완료 {fmt(ship_done)} / 잔량 {fmt(ship_left)} ({pct_ship:.1f}%)'
            )
        else:
            block_lines.append(f'   [납기]   -')

        sections.append('\n'.join(block_lines))

    # 전체 텍스트 합치기
    header_lines = ['<블룸>', '']
    if base_date:
        header_lines.append(f'[기준일] {base_date}')
        header_lines.append('')
    header_lines.append('[제품별 진행 현황]')
    header_lines.append('')

    raw_text = '\n'.join(header_lines) + '\n\n'.join(sections)

    status_hint = 'RED' if has_red else ('ORANGE' if has_orange else 'BLACK')

    return {
        'title': '<블룸>',
        'product_sections': sections,
        'raw_text': raw_text,
        'status_hint': status_hint,
    }


# ========================================
# 5. 3개 파일 한 번에 처리
# ========================================
def generate_bloom_note_from_files(file_paths: List[str]) -> Dict[str, Any]:
    """
    파일 3개를 받아서 자동 분류 후 노트 생성.
    파일명 무관, 어떤 순서로 와도 OK.
    """
    classified = {'po': None, 'bop_nct': None, 'kpe_nct': None}
    unknown = []

    for p in file_paths:
        kind = classify_excel(p)
        if kind == 'unknown':
            unknown.append(p)
        else:
            classified[kind] = p

    # 검증
    errors = []
    if not classified['po']:
        errors.append('실적 엑셀(출하계획+날짜별 생산실적 시트)을 찾을 수 없습니다.')
    # NCT는 선택사항

    if errors:
        return {
            'ok': False,
            'errors': errors,
            'classified': classified,
            'unknown_files': unknown,
        }

    # 파싱
    po_data = parse_po_excel(classified['po'])
    bop_nct = parse_nct_excel(classified['bop_nct']) if classified['bop_nct'] else None
    kpe_nct = parse_nct_excel(classified['kpe_nct']) if classified['kpe_nct'] else None

    note = build_bloom_note(po_data, bop_nct, kpe_nct)

    return {
        'ok': True,
        'classified': {k: Path(v).name if v else None for k, v in classified.items()},
        'unknown_files': [Path(p).name for p in unknown],
        'note': note,
        'po_data': po_data,
        'bop_nct_summary': {'model': bop_nct['model'], 'shortage': bop_nct['shortage_count'], 'total': bop_nct['total']} if bop_nct else None,
        'kpe_nct_summary': {'model': kpe_nct['model'], 'shortage': kpe_nct['shortage_count'], 'total': kpe_nct['total']} if kpe_nct else None,
    }



# ========================================
# 7. sections 구조 생성 (앱 호환 포맷)
# ========================================
def build_product_sections(
    po_data: Dict[str, Dict[str, Any]],
    bop_nct: Optional[Dict[str, Any]] = None,
    kpe_nct: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    """
    "제품별 진행 현황" 섹션의 items 리스트를 자동 생성한다.
    기존 sub + highlight 구조를 그대로 따라가서 앱이 변경 없이 인식할 수 있도록 한다.
    """
    material_map = {}
    if bop_nct and bop_nct.get('model'):
        material_map['BOP Assy'] = bop_nct
    if kpe_nct and kpe_nct.get('model'):
        material_map['KPE CORVA'] = kpe_nct

    products = [p for p in DEFAULT_PRODUCT_ORDER if p in po_data]
    for p in po_data.keys():
        if p not in products:
            products.append(p)

    items: List[Dict[str, Any]] = []

    for i, p in enumerate(products, 1):
        d = po_data.get(p, {})
        po = d.get('po', 0)
        prod = d.get('prod', 0)
        ship_done = d.get('ship_done', 0)
        ship_left = d.get('ship_left', 0)

        pct_prod = (prod / po * 100) if po else 0
        pct_ship = (ship_done / po * 100) if po else 0

        # sub: 제품명
        items.append({'type': 'sub', 'text': f'{i}) {p}'})

        # [원자재]
        nct = material_map.get(p)
        if nct and nct['shortage_count'] > 0:
            sc = nct['shortage_count']
            total = nct['total']
            color = '��' if sc >= 4 else '🟡'
            top_names = ', '.join(it['name'] for it in nct['top3'])
            text = f'[원자재] {color} NCT 자재 부족 {sc}/{total}종 ({top_names}{" 외 " + str(sc-3) + "종" if sc > 3 else ""})'
            items.append({'type': 'highlight', 'text': text})
        elif po == 0:
            items.append({'type': 'highlight', 'text': '[원자재]'})
        elif po > 0 and prod == 0 and ship_done == 0:
            items.append({'type': 'highlight', 'text': '[원자재] 미발주'})
        else:
            items.append({'type': 'highlight', 'text': '[원자재] 자재 점검 대상 없음'})

        # [입고]
        items.append({'type': 'highlight', 'text': '[입고]'})

        # [생산]
        if po:
            items.append({'type': 'highlight', 'text': f'[생산] {fmt(prod)} / {fmt(po)} 조립완료 ({pct_prod:.1f}%)'})
        else:
            items.append({'type': 'highlight', 'text': f'[생산] {fmt(prod)} (총 PO 정보 없음)'})

        # [납기]
        if po:
            items.append({'type': 'highlight',
                          'text': f'[납기] 총 PO {fmt(po)} / 출하완료 {fmt(ship_done)} / 잔량 {fmt(ship_left)} ({pct_ship:.1f}%)'})
        else:
            items.append({'type': 'highlight', 'text': '[납기]'})

    return items


def update_card_sections(existing_card: Dict[str, Any], product_section_items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    기존 카드의 sections 중 "제품별 진행 현황" 만 자동 갱신.
    이슈사항 등 다른 섹션은 그대로 둔다.
    """
    if not isinstance(existing_card, dict):
        existing_card = {}
    sections = existing_card.get('sections') or []
    if not isinstance(sections, list):
        sections = []

    target_title = '제품별 진행 현황'
    found = False
    for sec in sections:
        if isinstance(sec, dict) and sec.get('title') == target_title:
            sec['items'] = product_section_items
            found = True
            break
    if not found:
        sections.append({'title': target_title, 'items': product_section_items})

    existing_card['sections'] = sections
    return existing_card

# ========================================
# 6. CLI 테스트
# ========================================
if __name__ == '__main__':
    import sys
    import json

    if len(sys.argv) < 2:
        # 기본: test_data 폴더의 모든 xlsx
        test_dir = Path(__file__).parent / 'test_data'
        files = sorted([str(p) for p in test_dir.glob('*.xlsx')])
    else:
        files = sys.argv[1:]

    print(f'[INPUT] {len(files)}개 파일:')
    for f in files:
        print(f'  - {Path(f).name}')

    result = generate_bloom_note_from_files(files)

    print('\n[CLASSIFY]')
    for k, v in result.get('classified', {}).items():
        print(f'  {k}: {v}')

    if result.get('unknown_files'):
        print(f'\n[UNKNOWN] {result["unknown_files"]}')

    if not result['ok']:
        print('\n[ERRORS]')
        for e in result['errors']:
            print(f'  ! {e}')
        sys.exit(1)

    print('\n[NCT SUMMARY]')
    print(f'  BOP: {result.get("bop_nct_summary")}')
    print(f'  KPE: {result.get("kpe_nct_summary")}')

    print('\n' + '=' * 60)
    print('[ 자동 생성된 블룸 노트 ]')
    print('=' * 60)
    print(result['note']['raw_text'])
    print('\n[STATUS HINT]', result['note']['status_hint'])


# =========================================================
# === bloom-card-builder v2 ===
# generate_bloom_note_from_files() 결과의 note["raw_text"] 를 파싱해서
# 프론트(static/bloom_auto.js)가 그대로 쓰는 card 구조를 만든다.
# =========================================================
import re as _re

_BC_TAGS = ('원자재', '입고', '생산', '납기')

def _bc_pct(text):
    if not text:
        return None
    m = _re.search(r'(\d+(?:\.\d+)?)\s*%', text)
    return float(m.group(1)) if m else None

def _bc_status(label, *, pct=None):
    if not label or label.strip() in ('-', ''):
        return 'gray'
    s = label.strip()
    if ('부족' in s) or ('🔴' in s):
        return 'red'
    if pct is None:
        pct = _bc_pct(s)
    if pct is None:
        # 숫자 없는 안내 문구 (예: "자재 점검 대상 없음")
        return 'gray'
    if pct >= 100:
        return 'green'
    if pct >= 1:
        return 'yellow'
    return 'gray'

def _bc_step(label):
    """라벨에서 분자/분모/퍼센트를 분리해 step 객체로 반환."""
    if label is None:
        return {'status': 'gray', 'pct': 0, 'num': None, 'den': None,
                'fraction': '-', 'label': '-'}
    label = label.strip()
    if not label or label == '-':
        return {'status': 'gray', 'pct': 0, 'num': None, 'den': None,
                'fraction': '-', 'label': '-'}

    pct = _bc_pct(label)

    # 분자/분모 추출 — 납기/생산/원자재 라벨별 패턴 우선 적용
    import re as _re_local
    num = den = None

    # 1) 납기: "총 PO 10,419 / 출하완료 0" → den=PO, num=출하완료
    m_dlv = _re_local.search(
        r'총\s*PO\s*(\d[\d,]*)\s*/\s*출하완료\s*(\d[\d,]*)', label)
    if m_dlv:
        den = m_dlv.group(1).replace(',', '')
        num = m_dlv.group(2).replace(',', '')

    # 2) 일반 "N / M" 패턴 (생산, 원자재 부족 N/M종 포함)
    if num is None or den is None:
        m_gen = _re_local.search(r'(\d[\d,]*)\s*/\s*(\d[\d,]*)', label)
        if m_gen:
            num = m_gen.group(1).replace(',', '')
            den = m_gen.group(2).replace(',', '')

    has_shortage = ('부족' in label) or ('🔴' in label)
    status = _bc_status(label, pct=pct)
    if has_shortage:
        status = 'red'

    if num is not None and den is not None:
        # 분모에 콤마 다시 넣어서 보기 좋게
        try:
            fraction = "{:,} / {:,}".format(int(num), int(den))
        except Exception:
            fraction = "{} / {}".format(num, den)
    else:
        fraction = '-'

    return {
        'status': status,
        'pct': round(pct, 1) if pct is not None else 0,
        'num': num,
        'den': den,
        'fraction': fraction,
        'label': label,
    }

def _bc_parse_raw_text(raw):
    """raw_text → {'issues':[...], 'products':[{name, 원자재, 입고, 생산, 납기}, ...]}"""
    issues = []
    products = []
    if not raw:
        return issues, products

    lines = [ln.rstrip() for ln in raw.splitlines()]

    section = None  # '이슈사항' | '제품별 진행 현황' | None
    current = None

    re_header = _re.compile(r'^\[(.+?)\]\s*$')
    re_product_num = _re.compile(r'^\s*(\d+)\)\s*(.+?)\s*$')
    re_tag = _re.compile(r'^\s*\[(원자재|입고|생산|납기)\]\s*(.*)$')

    for ln in lines:
        if not ln.strip():
            continue

        mh = re_header.match(ln.strip())
        if mh:
            # 새 섹션 시작 시 진행중 제품 flush
            if current is not None:
                products.append(current)
                current = None
            section = mh.group(1).strip()
            continue

        if section == '이슈사항':
            txt = ln.strip().lstrip('-').strip()
            if txt:
                issues.append(txt)
            continue

        if section == '제품별 진행 현황':
            mp = re_product_num.match(ln)
            if mp:
                if current is not None:
                    products.append(current)
                current = {'name': mp.group(2).strip(), '_tags': {}}
                continue

            mt = re_tag.match(ln)
            if mt and current is not None:
                tag, val = mt.group(1), mt.group(2).strip()
                current['_tags'][tag] = val
                continue

    if current is not None:
        products.append(current)

    out = []
    for p in products:
        tags = p.get('_tags', {})
        out.append({
            'name': p.get('name', ''),
            'material': _bc_step(tags.get('원자재', '')),
            'inbound': _bc_step(tags.get('입고', '')),
            'production': _bc_step(tags.get('생산', '')),
            'delivery': _bc_step(tags.get('납기', '')),
            'material_detail': '',
        })
    return issues, out

def build_card_from_note(note: dict) -> dict:
    if not isinstance(note, dict):
        return {'title': '<블룸>', 'badge': '', 'issues': [], 'products': []}

    raw = note.get('raw_text') or ''
    issues, products = _bc_parse_raw_text(raw)

    # raw_text 에 이슈사항 섹션이 없는 경우 (현재 흐름) → 비워두기
    status_hint = (note.get('status_hint') or note.get('status') or '').upper()
    badge = 'RISK' if status_hint == 'RED' else ('주의' if status_hint == 'YELLOW' else '')

    return {
        'title': note.get('title') or '<블룸>',
        'badge': badge,
        'issues': issues,
        'products': products,
    }
